#!/usr/bin/env python3
"""
根据 Excel 文件中的时间戳创建剪映草稿，将图片插入到指定时间位置
"""

import os
import sys
import argparse
import re
import time
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image
import pycapcut as cc
from pycapcut import trange, tim

# ============ 默认配置 ============
DEFAULT_EXCEL_FILE = r"your_excel_file.xlsx"
DEFAULT_DRAFT_DIR = r"your_draft_path"
DEFAULT_DRAFT_NAME = "图片序列草稿"
DEFAULT_RESOLUTION = (1920, 1080)
# =================================


def extract_images_from_xlsx(xlsx_file_path, output_folder="extracted_images"):
    """
    从Excel文件中提取所有图片，按照y轴位置（行号）排序

    Args:
        xlsx_file_path: Excel文件路径
        output_folder: 输出图片的文件夹

    Returns:
        str: 输出文件夹的绝对路径
    """
    # 创建输出文件夹
    if not os.path.isabs(output_folder):
        excel_dir = os.path.dirname(xlsx_file_path)
        output_folder = os.path.join(excel_dir, output_folder)

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    print(f"\n=== 提取 Excel 图片 ===")
    print(f"从 {xlsx_file_path} 提取图片到 {output_folder}")

    # 加载工作簿
    wb = load_workbook(filename=xlsx_file_path)

    image_count = 0

    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        print(f"处理工作表: {sheet_name}")
        ws = wb[sheet_name]

        if hasattr(ws, '_images'):
            images = ws._images
            print(f"  找到 {len(images)} 张图片")

            # 收集所有图片及其位置信息
            images_with_position = []

            for i, img in enumerate(images):
                try:
                    # 获取图片数据 - 检查多种可能的属性
                    image_data = None

                    # 检查可能的属性名
                    for attr in ['_data', 'data', '_bytes', 'bytes', '_blob', 'blob']:
                        if hasattr(img, attr):
                            val = getattr(img, attr)
                            # 如果是方法则调用
                            if callable(val):
                                val = val()
                            if val and isinstance(val, bytes):
                                image_data = val
                                break

                    if image_data is None:
                        print(f"  图片 {i+1}: 无法获取图片数据")
                        continue

                    # 获取图片位置（行号）
                    row = 0
                    col = 0
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        if hasattr(anchor, '_from'):
                            from_pos = anchor._from
                            row = from_pos.row
                            col = from_pos.col

                    # 将图片信息和位置保存到列表
                    images_with_position.append({
                        'data': image_data,
                        'row': row,
                        'col': col,
                        'index': i
                    })

                except Exception as e:
                    print(f"  处理图片 {i+1} 失败: {e}")

            # 按照y轴位置（行号）排序，行号越大越靠后
            # 如果行号相同，则按照列号排序
            images_with_position.sort(key=lambda x: (x['row'], x['col']))

            # 按排序后的顺序保存图片
            for idx, img_info in enumerate(images_with_position):
                try:
                    # 打开并保存图片
                    image = Image.open(BytesIO(img_info['data']))
                    image_filename = os.path.join(output_folder, f"{sheet_name}_image_{idx+1}.png")
                    image.save(image_filename)
                    print(f"  已保存: {os.path.basename(image_filename)} (原位置: 行{img_info['row']+1}, 列{img_info['col']+1})")
                    image_count += 1
                except Exception as e:
                    print(f"  保存图片 {idx+1} 失败: {e}")

    print(f"✓ 共提取 {image_count} 张图片\n")
    return output_folder


def read_excel_timestamps(excel_path, sheet_name="Sheet1", column=1):
    """
    从 Excel 文件的第一列读取时间戳（保持原始行顺序）

    Args:
        excel_path: Excel 文件路径
        sheet_name: 工作表名称，默认 "Sheet1"
        column: 要读取的列号（1为第一列），默认为1

    Returns:
        List[Tuple[int, float]]: (行号, 时间戳) 的列表，按行号顺序
    """
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    timestamps = []
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row, column).value
        if cell_value is not None:
            try:
                timestamp = float(cell_value)
                if timestamp >= 0:
                    timestamps.append((row, timestamp))
            except (ValueError, TypeError):
                print(f"警告: 跳过无效值 {cell_value} (行 {row})")

    print(f"从 {excel_path} 读取了 {len(timestamps)} 个时间戳")
    print(f"时间戳（按Excel行顺序）:")
    for row, timestamp in timestamps:
        print(f"  行{row}: {timestamp}s")

    return timestamps


def collect_image_paths(images_folder, sheet_name="Sheet1"):
    """
    从文件夹中收集按数字排序的图片路径

    Args:
        images_folder: 图片文件夹路径
        sheet_name: 工作表名称，用于匹配文件名模式

    Returns:
        List[str]: 排序后的图片绝对路径列表
    """
    if not os.path.exists(images_folder):
        raise FileNotFoundError(f"图片文件夹不存在: {images_folder}")

    # 匹配模式: Sheet1_image_1.png, Sheet1_image_2.png, etc.
    pattern = re.compile(re.escape(sheet_name) + r'_image_(\d+)\.png', re.IGNORECASE)

    images = []
    for filename in os.listdir(images_folder):
        match = pattern.match(filename)
        if match:
            number = int(match.group(1))
            full_path = os.path.join(images_folder, filename)
            images.append((number, full_path))

    # 按数字排序
    images.sort(key=lambda x: x[0])

    image_paths = [path for _, path in images]

    print(f"找到 {len(image_paths)} 张图片")
    for i, path in enumerate(image_paths[:5], 1):
        print(f"  {i}. {os.path.basename(path)}")
    if len(image_paths) > 5:
        print(f"  ... 还有 {len(image_paths) - 5} 张")

    return image_paths


def check_timestamps_info(timestamps_with_rows, duration):
    """
    显示时间戳信息，不进行任何调整

    Args:
        timestamps_with_rows: (行号, 时间戳) 的列表
        duration: 每个片段的持续时间（秒）

    Returns:
        List[Tuple[int, float]]: 原始的 (行号, 时间戳) 列表
    """
    print(f"时间片段信息（每张图片显示 {duration} 秒）:")
    for row, timestamp in timestamps_with_rows:
        start = timestamp
        end = timestamp + duration
        print(f"  行{row}: {start}s - {end}s")

    return timestamps_with_rows


def create_capcut_draft(excel_path, images_folder, draft_dir,
                        draft_name, resolution=(1920, 1080)):
    """
    主函数：创建剪映草稿并插入图片

    Args:
        excel_path: Excel 文件路径
        images_folder: 提取的图片文件夹路径
        draft_dir: 剪映草稿目录
        draft_name: 草稿名称
        resolution: 视频分辨率 (宽, 高)，默认 (1920, 1080)
    """
    # 1. 读取时间戳（Excel中的数字代表显示时长）
    print(f"\n=== 读取 Excel 文件 ===")
    timestamps = read_excel_timestamps(excel_path)

    if not timestamps:
        raise ValueError("Excel 文件中没有找到有效的时间戳")

    # 2. 收集图片路径
    print(f"\n=== 收集图片文件 ===")
    image_paths = collect_image_paths(images_folder)

    if not image_paths:
        raise ValueError("图片文件夹中没有找到图片")

    # 3. 验证数量
    if len(image_paths) < len(timestamps):
        raise ValueError(
            f"图片数量不足: 找到 {len(image_paths)} 张图片，"
            f"但需要 {len(timestamps)} 张（对应 {len(timestamps)} 个时间戳）"
        )

    # 4. 计算每个图片的起始时间（按顺序叠加）
    print(f"\n=== 计算图片序列时间 ===")
    current_time = 0.0
    segments_info = []
    for i, ((row, timestamp), image_path) in enumerate(zip(timestamps, image_paths), 1):
        # Excel中的数字作为这张图片的显示时长
        duration = timestamp
        start_time = current_time
        end_time = current_time + duration
        segments_info.append({
            'index': i,
            'row': row,
            'image_path': image_path,
            'duration': duration,
            'start_time': start_time,
            'end_time': end_time
        })
        print(f"  片段 {i}: {start_time}s - {end_time}s (时长: {duration}s)")
        current_time = end_time
    print(f"✓ 总时长: {current_time}s")

    # 5. 创建剪映草稿
    print(f"\n=== 创建剪映草稿 ===")
    print(f"草稿目录: {draft_dir}")
    print(f"草稿名称: {draft_name}")
    print(f"分辨率: {resolution[0]}x{resolution[1]}")

    draft_folder = cc.DraftFolder(draft_dir)
    script = draft_folder.create_draft(
        draft_name,
        resolution[0],
        resolution[1],
        allow_replace=True
    )

    # 6. 添加单个视频轨道
    print(f"\n=== 添加视频轨道 ===")
    script.add_track(cc.TrackType.video, track_name="图片序列")
    print(f"✓ 已添加 1 个视频轨道")

    # 7. 按顺序插入图片到同一个轨道
    print(f"\n=== 插入图片序列 ===")
    for seg in segments_info:
        # 使用默认 ClipSettings（全屏显示）
        video_segment = cc.VideoSegment(
            seg['image_path'],
            trange(f"{seg['start_time']}s", f"{seg['duration']}s")
        )

        # 添加到同一个轨道
        script.add_segment(video_segment, track_name="图片序列")
        print(f"✓ 片段 {seg['index']}: {os.path.basename(seg['image_path'])} (Excel行{seg['row']}) @ {seg['start_time']}s - {seg['end_time']}s (时长: {seg['duration']}s)")

    # 8. 保存草稿
    print(f"\n=== 保存草稿 ===")
    script.save()
    print(f"✓ 草稿已保存: {draft_name}")

    print(f"\n=== 完成 ===")
    print(f"成功创建草稿 '{draft_name}'，包含 {len(segments_info)} 个图片片段，总时长 {current_time}s")
    print(f"打开剪映专业版即可查看该草稿")

    return script


def main():
    # 记录开始时间
    start_time = time.time()

    parser = argparse.ArgumentParser(
        description='根据 Excel 文件中的时间戳创建剪映草稿，将图片插入到指定时间位置',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument(
        '--excel',
        default=DEFAULT_EXCEL_FILE,
        help=f'Excel 文件路径（第一列包含时间戳，单位：秒），默认: {DEFAULT_EXCEL_FILE}'
    )

    parser.add_argument(
        '--images',
        help='提取的图片文件夹路径（默认为 Excel 文件同目录下的 extracted_images 文件夹）'
    )

    parser.add_argument(
        '--draft-dir',
        default=DEFAULT_DRAFT_DIR,
        help=f'剪映草稿目录（默认: {DEFAULT_DRAFT_DIR}）'
    )

    parser.add_argument(
        '--draft-name',
        default=DEFAULT_DRAFT_NAME,
        help=f'草稿名称，默认: {DEFAULT_DRAFT_NAME}'
    )


    parser.add_argument(
        '--resolution',
        nargs=2,
        type=int,
        default=DEFAULT_RESOLUTION,
        metavar=('WIDTH', 'HEIGHT'),
        help=f'视频分辨率（宽 高），默认: {DEFAULT_RESOLUTION[0]} {DEFAULT_RESOLUTION[1]}'
    )

    args = parser.parse_args()

    # 自动提取图片（如果图片文件夹不存在或为空）
    if args.images:
        images_folder = args.images
    else:
        # 获取程序所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        images_folder = os.path.join(script_dir, "extracted_images")

    # 检查是否需要提取图片
    need_extract = False
    if not os.path.exists(images_folder):
        need_extract = True
    else:
        # 检查文件夹是否为空
        if not os.listdir(images_folder):
            need_extract = True

    if need_extract:
        images_folder = extract_images_from_xlsx(args.excel, images_folder)

    try:
        create_capcut_draft(
            excel_path=args.excel,
            images_folder=images_folder,
            draft_dir=args.draft_dir,
            draft_name=args.draft_name,
            resolution=tuple(args.resolution)
        )

        # 计算并显示总时长
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"\n⏱ 任务完成！总耗时: {elapsed_time:.2f} 秒")

    except Exception as e:
        print(f"\n错误: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
